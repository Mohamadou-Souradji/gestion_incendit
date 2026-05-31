import json
import re
from collections import defaultdict

from openpyxl import load_workbook


def main() -> None:
    path = r"c:\Users\abous\Downloads\remonté d'incident opérationnel-12.xlsx"
    wb = load_workbook(path, data_only=False)

    result: dict = {
        "path": path,
        "sheets": [],
        "defined_names": [],
    }

    # Defined names (named ranges / constants)
    try:
        for dn in wb.defined_names.definedName:
            result["defined_names"].append(
                {
                    "name": dn.name,
                    "attr_text": dn.attr_text,
                    "comment": dn.comment,
                    "localSheetId": dn.localSheetId,
                }
            )
    except Exception:
        # Some workbooks may not expose defined names cleanly
        pass

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Used range
        dimension = ws.calculate_dimension()

        nonempty_cells = 0
        formula_cells = []

        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                nonempty_cells += 1
                if isinstance(v, str) and v.startswith("="):
                    formula_cells.append({"cell": c.coordinate, "formula": v})

        # Data validations
        validations = []
        if ws.data_validations is not None:
            for dv in ws.data_validations.dataValidation:
                validations.append(
                    {
                        "type": dv.type,
                        "operator": dv.operator,
                        "formula1": dv.formula1,
                        "formula2": dv.formula2,
                        "sqref": str(dv.sqref),
                        "allow_blank": dv.allowBlank,
                        "show_error": dv.showErrorMessage,
                        "error": dv.error,
                        "prompt": dv.prompt,
                    }
                )

        # Tables (structured ranges)
        tables = []
        for tbl in getattr(ws, "tables", {}).values():
            tables.append(
                {
                    "name": tbl.name,
                    "ref": tbl.ref,
                    "displayName": getattr(tbl, "displayName", None),
                }
            )

        # Simple heuristics for "inputs":
        # - cells covered by validation sqref
        dv_cells = set()
        for dv in validations:
            # sqref can be 'A1 A3' or 'A1:A10'
            for token in str(dv.get("sqref") or "").split():
                dv_cells.add(token)

        sheet_obj = {
            "name": sheet_name,
            "dimension": dimension,
            "nonempty_cells": nonempty_cells,
            "formula_count": len(formula_cells),
            "formula_samples": formula_cells[:50],
            "data_validation_count": len(validations),
            "data_validations": validations[:100],
            "tables": tables,
            "merged_cell_ranges": [str(rng) for rng in ws.merged_cells.ranges][:200],
        }

        result["sheets"].append(sheet_obj)

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

