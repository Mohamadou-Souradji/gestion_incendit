import json
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from typing import Optional, List, Dict, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter


def _strip_ns(tag: str) -> str:
    return tag.split("}", 1)[1] if "}" in tag else tag


def _iter_x14_validations(sheet_xml: str) -> List[dict]:
    root = ET.fromstring(sheet_xml.encode("utf-8"))
    vals = []
    for extlst in root.iter():
        if _strip_ns(extlst.tag) != "extLst":
            continue
        for ext in list(extlst):
            for node in ext.iter():
                if _strip_ns(node.tag) != "dataValidations":
                    continue
                for dv in list(node):
                    if _strip_ns(dv.tag) != "dataValidation":
                        continue
                    # formula is inside x14:formula1/xm:f and sqref inside xm:sqref
                    formula = None
                    sqref = None
                    for ch in dv.iter():
                        local = _strip_ns(ch.tag)
                        if local == "f" and (ch.text or "").strip():
                            formula = (ch.text or "").strip()
                        elif local == "sqref" and (ch.text or "").strip():
                            sqref = (ch.text or "").strip()
                    vals.append(
                        {
                            "type": dv.attrib.get("type"),
                            "allowBlank": dv.attrib.get("allowBlank"),
                            "sqref": sqref,
                            "formula": formula,
                        }
                    )
    return vals


def _expand_sqref_tokens(sqref: str) -> List[str]:
    # returns list of range tokens (e.g. ["F39:H39", "J12"])
    return [t.strip() for t in (sqref or "").split() if t.strip()]


def _range_top_left(token: str) -> str:
    # "F39:H39" -> "F39"; "J12" -> "J12"
    return token.split(":", 1)[0]


def main() -> None:
    xlsx_path = r"c:\Users\abous\Downloads\remonté d'incident opérationnel-12.xlsx"

    # Load workbook values for label extraction + list values
    wb = load_workbook(xlsx_path, data_only=True)
    ws_form = wb["Fiche de remontée "]
    ws_lists = wb["datavalidation"]

    # Read x14 validations from sheet1.xml
    with zipfile.ZipFile(xlsx_path, "r") as z:
        sheet1_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="replace")
    dvs = _iter_x14_validations(sheet1_xml)

    # Build list values for each formula like "datavalidation!$B$2:$B$61"
    def read_list(formula: str) -> List[str]:
        if not formula:
            return []
        # formats: datavalidation!$B$2:$B$61 OR datavalidation!C2:C9
        m = formula.split("!", 1)
        if len(m) != 2:
            return []
        rng = m[1].replace("$", "")
        if ":" not in rng:
            # single cell
            v = ws_lists[rng].value
            return [str(v).strip()] if v is not None and str(v).strip() else []
        start, end = rng.split(":", 1)
        sc, sr = coordinate_from_string(start)
        ec, er = coordinate_from_string(end)
        s_col = column_index_from_string(sc)
        e_col = column_index_from_string(ec)
        vals = []
        for r in range(sr, er + 1):
            for c in range(s_col, e_col + 1):
                cell = f"{get_column_letter(c)}{r}"
                v = ws_lists[cell].value
                if v is None:
                    continue
                sv = str(v).strip()
                if sv:
                    vals.append(sv)
        return vals

    # Try to find a label for a given input cell by scanning left on same row
    def find_label(input_cell: str) -> Optional[str]:
        col_letter, row = coordinate_from_string(input_cell)
        col = column_index_from_string(col_letter)
        for c in range(col - 1, 0, -1):
            v = ws_form.cell(row=row, column=c).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                return s
        return None

    fields = []
    for dv in dvs:
        tokens = _expand_sqref_tokens(dv["sqref"] or "")
        if not tokens:
            continue
        choices = read_list(dv["formula"])
        for token in tokens:
            top_left = _range_top_left(token)
            label = find_label(top_left)
            fields.append(
                {
                    "sqref": token,
                    "top_left": top_left,
                    "label_guess": label,
                    "validation_type": dv.get("type"),
                    "formula": dv.get("formula"),
                    "choices": choices,
                }
            )

    out = {
        "xlsx_path": xlsx_path,
        "sheet": "Fiche de remontée ",
        "fields_from_validations": fields,
    }
    with open("excel_fields.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print("Wrote excel_fields.json with", len(fields), "field mappings")


if __name__ == "__main__":
    main()

