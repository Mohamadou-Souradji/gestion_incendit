import csv
import datetime
import io
from decimal import Decimal

from django.contrib import messages
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone

from .forms import (
    ActionCorrectiveFormSet,
    AffectationAgentForm,
    AvisDirRisquesForm,
    AvisValidationFormSet,
    IncidentForm,
    IncidentImportForm,
    MesureImmediateFormSet,
    ValidationChefForm,
)
from .models import AvisValidation, Incident

EXPORT_FIELDS = [
    "id","version","date_declaration","date_mise_a_jour","agence","region","direction",
    "nom_prenom_declarant","fonction","poste_telephonique","email","date_debut_incident",
    "date_decouverte_incident","description","incident_lie_risque_credit","categorisation_baloise",
    "sous_categorie_incident","macro_processus","processus","domaine_activite","incident_si",
    "incident_non_conformite","montant_estime_perte","comptabilisation_perte",
    "date_comptabilisation_perte","montant_recuperations","date_recuperation","nature_recuperation",
    "montant_net_perte","criticite","impact_pca","statut_incident","date_butoire_resolution",
    "commentaires","validation_chef","created_at","updated_at",
]


def home(request):
    return redirect("incidents:list")


def _profile(request):
    try:
        return request.user.profile
    except Exception:
        return None


def _to_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _to_decimal(value):
    if value in (None, ""):
        return None
    raw = str(value).replace(" ", "").replace(",", ".")
    try:
        return Decimal(raw)
    except Exception:
        return None


def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _incidents_for_user(request):
    """Filtre le queryset selon le rôle connecté."""
    p = _profile(request)
    if p is None:
        return Incident.objects.none()

    qs = Incident.objects.select_related("declarant","agent_affecte").all()

    if p.is_declarant:
        # Uniquement ses propres déclarations
        return qs.filter(declarant=request.user)

    if p.is_chef:
        # Tous les incidents de sa direction (en attente + traités)
        return qs.filter(direction=p.direction)

    if p.is_risques_op:
        # Uniquement les incidents qui lui sont affectés
        return qs.filter(agent_affecte=request.user)

    # DIR_RISQUES et ADMIN voient tout
    return qs


def _filtered_incidents(request, base_qs=None):
    if base_qs is None:
        base_qs = Incident.objects.all()
    criticite  = request.GET.get("criticite", "").strip()
    statut     = request.GET.get("statut", "").strip()
    direction  = request.GET.get("direction", "").strip()
    date_from  = _to_date(request.GET.get("date_from", ""))
    date_to    = _to_date(request.GET.get("date_to", ""))
    if criticite: base_qs = base_qs.filter(criticite=criticite)
    if statut:    base_qs = base_qs.filter(statut_incident=statut)
    if direction: base_qs = base_qs.filter(direction__icontains=direction)
    if date_from: base_qs = base_qs.filter(date_declaration__gte=date_from)
    if date_to:   base_qs = base_qs.filter(date_declaration__lte=date_to)
    return base_qs


# ─── Liste ──────────────────────────────────────────────────────────────────

@login_required
def incident_list(request):
    p = _profile(request)
    base_qs   = _incidents_for_user(request)
    incidents = _filtered_incidents(request, base_qs)[:200]
    return render(request, "incidents/incident_list.html", {
        "incidents": incidents,
        "profile": p,
        "filter_criticite":  request.GET.get("criticite", ""),
        "filter_statut":     request.GET.get("statut", ""),
        "filter_direction":  request.GET.get("direction", ""),
        "filter_date_from":  request.GET.get("date_from", ""),
        "filter_date_to":    request.GET.get("date_to", ""),
    })


# ─── Créer ──────────────────────────────────────────────────────────────────

@login_required
def incident_create(request):
    p = _profile(request)
    if p and not (p.is_declarant or p.is_admin):
        messages.error(request, "Seul un déclarant peut créer un incident.")
        return redirect("incidents:list")

    incident = Incident()
    if request.method == "POST":
        form       = IncidentForm(request.POST, instance=incident)
        mesures_fs = MesureImmediateFormSet(request.POST, instance=incident, prefix="mesures")
        actions_fs = ActionCorrectiveFormSet(request.POST, instance=incident, prefix="actions")
        if form.is_valid() and mesures_fs.is_valid() and actions_fs.is_valid():
            incident = form.save(commit=False)
            incident.declarant = request.user
            # Date de déclaration = aujourd'hui si non renseignée
            if not incident.date_declaration:
                incident.date_declaration = timezone.localdate()
            if p and p.direction and not incident.direction:
                incident.direction = p.direction
            incident.save()
            mesures_fs.instance = incident
            actions_fs.instance = incident
            mesures_fs.save()
            actions_fs.save()
            messages.success(request, f"Incident #{incident.pk} déclaré — en attente de validation par votre chef.")
            return redirect("incidents:detail", pk=incident.pk)
    else:
        initial = {"date_declaration": timezone.localdate()}
        if p and p.direction:
            initial["direction"] = p.direction
        if request.user.get_full_name():
            initial["nom_prenom_declarant"] = request.user.get_full_name()
        if request.user.email:
            initial["email"] = request.user.email
        form       = IncidentForm(instance=incident, initial=initial)
        mesures_fs = MesureImmediateFormSet(instance=incident, prefix="mesures")
        actions_fs = ActionCorrectiveFormSet(instance=incident, prefix="actions")

    return render(request, "incidents/incident_form.html", {
        "form": form, "mesures_formset": mesures_fs, "actions_formset": actions_fs,
        "mode": "create", "profile": p,
    })


# ─── Détail ─────────────────────────────────────────────────────────────────

@login_required
def incident_detail(request, pk):
    p = _profile(request)
    incident = get_object_or_404(Incident, pk=pk)

    # Contrôle d'accès
    if p:
        if p.is_declarant and incident.declarant != request.user:
            messages.error(request, "Accès refusé.")
            return redirect("incidents:list")
        if p.is_chef and incident.direction != p.direction:
            messages.error(request, "Accès refusé.")
            return redirect("incidents:list")
        if p.is_risques_op and incident.agent_affecte != request.user:
            messages.error(request, "Cet incident ne vous est pas affecté.")
            return redirect("incidents:list")

    for role in (AvisValidation.ROLE_RISQUES_OP, AvisValidation.ROLE_DIR_GESTION_RISQUES):
        AvisValidation.objects.get_or_create(incident=incident, role=role)
    avis = {a.role: a for a in incident.avis_validations.all()}

    return render(request, "incidents/incident_detail.html", {
        "incident": incident,
        "mesures":  incident.mesures_immediates.all(),
        "actions":  incident.actions_correctives.all(),
        "avis_risques_op":  avis.get(AvisValidation.ROLE_RISQUES_OP),
        "avis_dir_risques": avis.get(AvisValidation.ROLE_DIR_GESTION_RISQUES),
        "profile": p,
    })


# ─── Modifier ───────────────────────────────────────────────────────────────

@login_required
def incident_edit(request, pk):
    p = _profile(request)
    incident = get_object_or_404(Incident, pk=pk)

    can_edit = (
        (p and p.is_declarant and incident.declarant == request.user and incident.validation_chef == "EN_ATTENTE")
        or (p and p.is_admin)
    )
    if not can_edit:
        messages.error(request, "Modification non autorisée.")
        return redirect("incidents:detail", pk=pk)

    if request.method == "POST":
        form       = IncidentForm(request.POST, instance=incident)
        mesures_fs = MesureImmediateFormSet(request.POST, instance=incident, prefix="mesures")
        actions_fs = ActionCorrectiveFormSet(request.POST, instance=incident, prefix="actions")
        avis_fs    = AvisValidationFormSet(request.POST, instance=incident, prefix="avis")
        if form.is_valid() and mesures_fs.is_valid() and actions_fs.is_valid() and avis_fs.is_valid():
            form.save(); mesures_fs.save(); actions_fs.save(); avis_fs.save()
            messages.success(request, "Incident mis à jour.")
            return redirect("incidents:detail", pk=incident.pk)
    else:
        form       = IncidentForm(instance=incident)
        mesures_fs = MesureImmediateFormSet(instance=incident, prefix="mesures")
        actions_fs = ActionCorrectiveFormSet(instance=incident, prefix="actions")
        for role in (AvisValidation.ROLE_RISQUES_OP, AvisValidation.ROLE_DIR_GESTION_RISQUES):
            AvisValidation.objects.get_or_create(incident=incident, role=role)
        avis_fs = AvisValidationFormSet(instance=incident, prefix="avis")

    return render(request, "incidents/incident_form.html", {
        "form": form, "mesures_formset": mesures_fs, "actions_formset": actions_fs,
        "avis_formset": avis_fs, "mode": "edit", "incident": incident, "profile": p,
    })


# ─── Validation chef ────────────────────────────────────────────────────────

@login_required
def chef_validation(request, pk):
    p = _profile(request)
    if not p or not p.is_chef:
        messages.error(request, "Réservé au chef de direction.")
        return redirect("incidents:list")

    # Le chef ne voit que les incidents de sa direction
    incident = get_object_or_404(Incident, pk=pk, direction=p.direction)

    if request.method == "POST":
        form = ValidationChefForm(request.POST, instance=incident)
        if form.is_valid():
            obj = form.save(commit=False)
            if not obj.date_validation_chef:
                obj.date_validation_chef = timezone.localdate()
            obj.save()
            messages.success(request, f"Incident #{pk} : {incident.get_validation_chef_display()}.")
            return redirect("incidents:list")
    else:
        form = ValidationChefForm(instance=incident, initial={"date_validation_chef": timezone.localdate()})

    return render(request, "incidents/chef_validation.html", {
        "incident": incident, "form": form, "profile": p,
    })


# ─── Tableau de bord risques ─────────────────────────────────────────────────

@login_required
def risques_dashboard(request):
    p = _profile(request)
    if not p or not p.can_treat:
        messages.error(request, "Accès réservé à la direction des risques.")
        return redirect("incidents:list")

    # DIR_RISQUES voit tous les incidents validés, peut affecter
    # RISQUES_OP ne devrait pas accéder directement (il voit sa liste)
    qs = Incident.objects.filter(validation_chef="VALIDE") \
        .select_related("declarant","agent_affecte") \
        .prefetch_related("avis_validations")[:300]

    return render(request, "incidents/risques_dashboard.html", {"incidents": qs, "profile": p})


# ─── Affectation agent ───────────────────────────────────────────────────────

@login_required
def affecter_agent(request, pk):
    p = _profile(request)
    if not p or not (p.is_dir_risques or p.is_admin):
        messages.error(request, "Réservé au directeur des risques.")
        return redirect("incidents:list")

    incident = get_object_or_404(Incident, pk=pk, validation_chef="VALIDE")

    if request.method == "POST":
        form = AffectationAgentForm(request.POST, instance=incident)
        if form.is_valid():
            form.save()
            messages.success(request, f"Incident #{pk} affecté à {incident.agent_affecte}.")
            return redirect("incidents:detail", pk=pk)
    else:
        form = AffectationAgentForm(instance=incident)

    return render(request, "incidents/affecter_agent.html", {
        "incident": incident, "form": form, "profile": p,
    })


# ─── Avis direction risques ──────────────────────────────────────────────────

@login_required
def risques_avis(request, pk):
    p = _profile(request)
    if not p or not p.can_treat:
        messages.error(request, "Accès réservé.")
        return redirect("incidents:list")

    incident = get_object_or_404(Incident, pk=pk)

    # Agent : ne peut donner avis que sur incident qui lui est affecté
    if p.is_risques_op and incident.agent_affecte != request.user:
        messages.error(request, "Cet incident ne vous est pas affecté.")
        return redirect("incidents:list")

    avis, _ = AvisValidation.objects.get_or_create(
        incident=incident, role=AvisValidation.ROLE_DIR_GESTION_RISQUES
    )
    if request.method == "POST":
        form = AvisDirRisquesForm(request.POST, instance=avis)
        if form.is_valid():
            form.save()
            messages.success(request, "Avis enregistré.")
            return redirect("incidents:risques")
    else:
        form = AvisDirRisquesForm(instance=avis, initial={"date": timezone.localdate()})

    return render(request, "incidents/risques_avis.html", {
        "incident": incident, "form": form, "avis_dir": avis, "profile": p,
    })


# ─── Import ──────────────────────────────────────────────────────────────────

@login_required
def incident_import(request):
    p = _profile(request)
    if not p or not (p.is_declarant or p.is_admin):
        messages.error(request, "Accès refusé.")
        return redirect("incidents:list")

    if request.method == "POST":
        form = IncidentImportForm(request.POST, request.FILES)
        if form.is_valid():
            f    = form.cleaned_data["fichier"]
            mode = form.cleaned_data["mode"]
            try:
                if mode == IncidentImportForm.MODE_TEMPLATE:
                    inc = _import_from_template_excel(f)
                    inc.declarant = request.user
                    if not inc.date_declaration:
                        inc.date_declaration = timezone.localdate()
                    inc.save()
                    messages.success(request, f"Import réussi : incident #{inc.pk} créé.")
                    return redirect("incidents:detail", pk=inc.pk)
                created = _import_from_table(f, request.user)
                messages.success(request, f"Import terminé : {created} incident(s) créé(s).")
                return redirect("incidents:list")
            except Exception as exc:
                messages.error(request, f"Échec de l'import : {exc}")
    else:
        form = IncidentImportForm()

    return render(request, "incidents/incident_import.html", {"form": form, "profile": p})


# ─── Export ──────────────────────────────────────────────────────────────────

def _incident_row(incident):
    row = []
    for field in EXPORT_FIELDS:
        value = getattr(incident, field, "")
        if isinstance(value, datetime.datetime):
            row.append(value.strftime("%Y-%m-%d %H:%M:%S"))
        elif isinstance(value, datetime.date):
            row.append(value.strftime("%Y-%m-%d"))
        elif value is None:
            row.append("")
        else:
            row.append(str(value))
    return row


@login_required
def incident_export_csv(request):
    qs = _filtered_incidents(request, _incidents_for_user(request))
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="incidents_export.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(EXPORT_FIELDS)
    for inc in qs:
        writer.writerow(_incident_row(inc))
    return response


@login_required
def incident_export_xlsx(request):
    from openpyxl import Workbook
    qs = _filtered_incidents(request, _incidents_for_user(request))
    wb = Workbook(); ws = wb.active; ws.title = "incidents"
    ws.append(EXPORT_FIELDS)
    for inc in qs:
        ws.append(_incident_row(inc))
    out = io.BytesIO(); wb.save(out); out.seek(0)
    response = HttpResponse(out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="incidents_export.xlsx"'
    return response


@login_required
def incident_export_one_csv(request, pk):
    inc = get_object_or_404(Incident, pk=pk)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="incident_{pk}.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(EXPORT_FIELDS); writer.writerow(_incident_row(inc))
    return response


@login_required
def incident_export_one_xlsx(request, pk):
    from openpyxl import Workbook
    inc = get_object_or_404(Incident, pk=pk)
    wb = Workbook(); ws = wb.active; ws.title = f"incident_{pk}"
    ws.append(EXPORT_FIELDS); ws.append(_incident_row(inc))
    out = io.BytesIO(); wb.save(out); out.seek(0)
    response = HttpResponse(out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="incident_{pk}.xlsx"'
    return response


@login_required
def incident_export_one_pdf(request, pk):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    import os

    inc = get_object_or_404(Incident, pk=pk)
    font_name = "Helvetica"
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        for path in [r"C:\Windows\Fonts\arial.ttf", r"C:\Windows\Fonts\tahoma.ttf"]:
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont("CustomFont", path))
                font_name = "CustomFont"; break
    except Exception:
        pass

    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Heading2"], fontName=font_name, fontSize=14, leading=18)
    b = ParagraphStyle("b", parent=styles["BodyText"],  fontName=font_name, fontSize=10, leading=13)
    s = ParagraphStyle("s", parent=styles["BodyText"],  fontName=font_name, fontSize=9,  leading=11)

    def p(txt, style=b):
        safe = (txt or "").replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace("\n","<br/>")
        return Paragraph(safe, style)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    story = [
        p(f"Incident #{inc.pk}", h),
        p(f"Créé le {inc.created_at:%d/%m/%Y %H:%M} — mis à jour le {inc.updated_at:%d/%m/%Y %H:%M}", s),
        Spacer(1, 12),
    ]
    rows = [
        ["Direction",        inc.direction or "—"],
        ["Déclarant",        inc.nom_prenom_declarant or "—"],
        ["Date déclaration", inc.date_declaration.isoformat() if inc.date_declaration else "—"],
        ["Criticité",        inc.criticite or "—"],
        ["Statut",           inc.statut_incident or "—"],
        ["Validation chef",  inc.get_validation_chef_display()],
    ]
    t = Table(rows, colWidths=[170, 310])
    t.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,-1),font_name),("FONTSIZE",(0,0),(-1,-1),9),
        ("BACKGROUND",(0,0),(0,-1),colors.whitesmoke),
        ("GRID",(0,0),(-1,-1),0.5,colors.lightgrey),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
    ]))
    story.extend([t, Spacer(1,12), p("Description", h), p(inc.description or "—")])
    doc.build(story)
    response = HttpResponse(buf.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="incident_{pk}.pdf"'
    return response


# ─── Import helpers ──────────────────────────────────────────────────────────

def _import_from_template_excel(uploaded_file):
    from openpyxl import load_workbook
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb["Fiche de remontée "]
    data = {
        "version": _clean_text(ws["J12"].value),
        "date_declaration": _to_date(ws["J9"].value),
        "date_mise_a_jour": _to_date(ws["J11"].value),
        "agence": _clean_text(ws["F16"].value),
        "region": _clean_text(ws["F17"].value),
        "direction": _clean_text(ws["F18"].value),
        "nom_prenom_declarant": _clean_text(ws["F20"].value),
        "fonction": _clean_text(ws["F21"].value),
        "poste_telephonique": _clean_text(ws["F22"].value),
        "email": _clean_text(ws["F23"].value),
        "date_debut_incident": _to_date(ws["F28"].value),
        "date_decouverte_incident": _to_date(ws["F29"].value),
        "description": _clean_text(ws["C32"].value),
        "criticite": _clean_text(ws["F78"].value),
        "impact_pca": _clean_text(ws["F79"].value),
        "statut_incident": _clean_text(ws["F80"].value),
        "date_butoire_resolution": _to_date(ws["F81"].value),
        "commentaires": _clean_text(ws["F83"].value),
    }
    return Incident(**data)


TABLE_FIELD_MAP = {
    "version":"version","date_declaration":"date_declaration","date_mise_a_jour":"date_mise_a_jour",
    "agence":"agence","region":"region","direction":"direction",
    "nom_prenom_declarant":"nom_prenom_declarant","fonction":"fonction",
    "poste_telephonique":"poste_telephonique","email":"email",
    "date_debut_incident":"date_debut_incident","date_decouverte_incident":"date_decouverte_incident",
    "description":"description","incident_lie_risque_credit":"incident_lie_risque_credit",
    "categorisation_baloise":"categorisation_baloise","sous_categorie_incident":"sous_categorie_incident",
    "macro_processus":"macro_processus","processus":"processus",
    "domaine_activite":"domaine_activite","incident_si":"incident_si",
    "incident_non_conformite":"incident_non_conformite","montant_estime_perte":"montant_estime_perte",
    "comptabilisation_perte":"comptabilisation_perte","date_comptabilisation_perte":"date_comptabilisation_perte",
    "montant_recuperations":"montant_recuperations","date_recuperation":"date_recuperation",
    "nature_recuperation":"nature_recuperation","montant_net_perte":"montant_net_perte",
    "criticite":"criticite","impact_pca":"impact_pca","statut_incident":"statut_incident",
    "date_butoire_resolution":"date_butoire_resolution","commentaires":"commentaires",
}


def _normalize_header(h):
    return (h.lower()
        .replace("é","e").replace("è","e").replace("ê","e")
        .replace("à","a").replace("ù","u").replace("ô","o").replace("ï","i")
        .replace(" ","_").replace("-","_"))


def _import_from_table(uploaded_file, user=None):
    rows = []
    fn = uploaded_file.name.lower()
    if fn.endswith(".csv"):
        content = uploaded_file.read().decode("utf-8-sig", errors="replace")
        import csv as _csv
        rows = list(_csv.DictReader(io.StringIO(content)))
    else:
        from openpyxl import load_workbook
        wb = load_workbook(uploaded_file, data_only=True); ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

    created = 0
    for row in rows:
        mapped = {}
        for k, v in row.items():
            if not k: continue
            target = TABLE_FIELD_MAP.get(_normalize_header(str(k)))
            if not target: continue
            if target.startswith("date_"):    mapped[target] = _to_date(v)
            elif target.startswith("montant_"): mapped[target] = _to_decimal(v)
            else: mapped[target] = _clean_text(v)
        if mapped:
            if user: mapped["declarant"] = user
            if not mapped.get("date_declaration"):
                mapped["date_declaration"] = timezone.localdate()
            Incident.objects.create(**mapped)
            created += 1
    return created
