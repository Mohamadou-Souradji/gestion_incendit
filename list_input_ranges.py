import json

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, get_column_letter, coordinate_from_string, column_index_from_string


def find_label(ws, row: int, col: int):
    # scan left on same row
    for c in range(col - 1, 0, -1):
        v = ws.cell(row=row, column=c).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            return s
    return None


def main() -> None:
    xlsx_path = r"c:\Users\abous\Downloads\remonté d'incident opérationnel-12.xlsx"
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Fiche de remontée "]

    merged = list(ws.merged_cells.ranges)
    inputs = []
    for rng in merged:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))

        # we are interested in right-side input area (F=6 and beyond)
        if max_col < 6:
            continue

        # If there is already a value in the top-left, treat it as a header not input
        tl_val = ws.cell(row=min_row, column=min_col).value
        if tl_val is not None and str(tl_val).strip():
            continue

        label = find_label(ws, min_row, min_col)
        inputs.append(
            {
                "range": str(rng),
                "top_left": f"{get_column_letter(min_col)}{min_row}",
                "row": min_row,
                "col": min_col,
                "label_guess": label,
                "size": {"rows": max_row - min_row + 1, "cols": max_col - min_col + 1},
            }
        )

    inputs.sort(key=lambda x: (x["row"], x["col"]))
    out = {
        "xlsx_path": xlsx_path,
        "sheet": "Fiche de remontée ",
        "merged_input_ranges": inputs,
    }
    with open("excel_input_ranges.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print("Wrote excel_input_ranges.json with", len(inputs), "candidate inputs")


if __name__ == "__main__":
    main()

