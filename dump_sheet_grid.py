from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string


def main() -> None:
    xlsx_path = r"c:\Users\abous\Downloads\remonté d'incident opérationnel-12.xlsx"
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Fiche de remontée "]

    # Dump a grid around the missing sections
    start_col = column_index_from_string("B")
    end_col = column_index_from_string("M")
    start_row = 56
    end_row = 92

    for r in range(start_row, end_row + 1):
        parts = []
        for c in range(start_col, end_col + 1):
            coord = f"{get_column_letter(c)}{r}"
            v = ws.cell(row=r, column=c).value
            if v is None:
                s = ""
            else:
                s = str(v).replace("\n", " ").strip()
            # keep it compact but readable
            if len(s) > 40:
                s = s[:37] + "…"
            parts.append(f"{coord}={s}")
        line = " | ".join(parts)
        # only print rows that have something in them
        if any(p.split("=", 1)[1] for p in parts):
            print(line)


if __name__ == "__main__":
    main()

